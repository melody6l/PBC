"""Excel读写处理 - 读取清单文件、导出匹配结果"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def find_name_column(headers):
    """自动识别包含文件名称的列"""
    name_keywords = ["文件名", "名称", "资料名称", "文件名称", "文档名称", "清单名称", "材料名称"]
    for i, header in enumerate(headers):
        if isinstance(header, str):
            for kw in name_keywords:
                if kw in header:
                    return i
    # 默认返回第二列（通常序号在第一列，名称在第二列）
    return 1 if len(headers) > 1 else 0


def read_checklist(file_path):
    """读取Excel清单文件，返回列头和数据"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    if not rows:
        return {"headers": [], "data": [], "name_col_index": 0, "items": []}

    headers = rows[0]
    data = rows[1:]
    name_col_index = find_name_column(headers)

    items = []
    for row in data:
        if row and len(row) > name_col_index:
            name = row[name_col_index]
            if name and str(name).strip():
                items.append(str(name).strip())

    wb.close()
    return {
        "headers": [str(h) if h else "" for h in headers],
        "data": [[str(cell) if cell else "" for cell in row] for row in data],
        "name_col_index": name_col_index,
        "items": items,
    }


def export_results(results, headers, data, name_col_index):
    """将匹配结果导出为Excel文件，过滤空行空列"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文件核对结果"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # 过滤空列：跳过所有数据行中均为空的列
    valid_cols = []
    for col_idx in range(len(headers)):
        has_data = any(
            row[col_idx] and row[col_idx].strip()
            for row in data if len(row) > col_idx
        )
        if headers[col_idx].strip() or has_data:
            valid_cols.append(col_idx)

    # 过滤空行：跳过名称列为空或整行有效列都为空的行
    valid_rows = []
    for row in data:
        name_val = row[name_col_index] if len(row) > name_col_index and (name_col_index in valid_cols) else ""
        has_any = any(row[c] and row[c].strip() for c in valid_cols if len(row) > c)
        if has_any and name_val and name_val.strip():
            valid_rows.append(row)

    # 写入列头：有效原始列 + 核对结果(倒数第二列) + 文件超链接(最后一列)
    col_headers = [headers[c] for c in valid_cols] + ["核对结果", "文件超链接"]
    for i, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    row_idx = 2
    for result in results:
        # 找到对应的原始数据行
        for row in data:
            val = row[name_col_index] if len(row) > name_col_index else ""
            if val and val.strip() == result["checklist_name"]:
                orig_row = row
                break

        # 写入有效原始列数据
        if orig_row:
            for col_num, col_idx in enumerate(valid_cols, 1):
                cell_val = orig_row[col_idx] if len(orig_row) > col_idx else ""
                ws.cell(row=row_idx, column=col_num, value=cell_val)

        # 核对结果（倒数第二列）
        status_col = len(valid_cols) + 1
        status_cell = ws.cell(row=row_idx, column=status_col, value=result["status"])
        status_cell.alignment = Alignment(horizontal="center")
        if result["status"] == "已获取":
            status_cell.fill = green_fill
        else:
            status_cell.fill = red_fill

        # 文件超链接（最后一列）
        link_col = len(valid_cols) + 2
        matched_str = ", ".join(result["matched_names"]) if result["matched_names"] else ""
        ws.cell(row=row_idx, column=link_col, value=matched_str)

        if result["matched_files"]:
            link_path = result["matched_files"][0]
            link_url = "file:///" + link_path.replace("\\", "/")
            ws.cell(row=row_idx, column=link_col).hyperlink = link_url

        row_idx += 1

    # 设置列宽
    for i in range(1, len(col_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    import os
    os.makedirs("exports", exist_ok=True)
    output_path = "exports/result.xlsx"
    wb.save(output_path)
    wb.close()
    return output_path